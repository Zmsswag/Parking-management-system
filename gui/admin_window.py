# gui/admin_window.py
import sqlite3
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt5.QtWidgets import (QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, 
                             QMessageBox, QGroupBox, QDesktopWidget, QDialog, QDateEdit)
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import Qt, QTimer
import os

from core.parking_system import ParkingSystem
from core.plate_recognizer import PlateRecognizer
# from .login_window import LoginWindow  # <--- 删除此处的导入
from .dialogs import AddUserDialog, SearchDialog, BindVehicleDialog, DeleteUserDialog, MonthSelectionDialog

class AdminWindow(QWidget):
    def __init__(self, username):
        super().__init__()
        self.username = username
        self.parking = ParkingSystem(100)
        self.recognizer = PlateRecognizer()
        
        self.is_recognizing = False
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.login_window_instance = None # 用于持有新登录窗口的引用
        self.exit_window = None 
        
        self.init_ui()
        
    def init_ui(self):
        self.setMinimumSize(1200, 700)
        self.center()
        self.setWindowTitle(f'管理员控制台 - {self.username}')
        
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        left_widget = self.create_left_panel()
        right_widget = self.create_right_panel()
        
        main_layout.addWidget(left_widget, 2)
        main_layout.addWidget(right_widget, 8)
        
    def create_left_panel(self):
        """创建左侧的功能按钮面板"""
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setSpacing(15)
        
        button_style = """
            QPushButton {
                background-color: #34495e; color: white; border-radius: 5px;
                padding: 12px; font-size: 14px; text-align: left; padding-left: 20px;
            }
            QPushButton:hover { background-color: #4a6fa5; }
            QPushButton:pressed { background-color: #2c3e50; }
        """
        
        title_label = QLabel('功能菜单')
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; padding-bottom: 10px; border-bottom: 2px solid #bdc3c7;")
        title_label.setAlignment(Qt.AlignCenter)
        left_layout.addWidget(title_label)
        
        buttons_config = [
            (' 添加用户', self.show_add_user_dialog),
            (' 删除用户', self.show_delete_user_dialog),
            (' 绑定车辆', self.show_bind_vehicle_dialog),
            (' 查询记录', self.show_search_dialog),
            (' 收费统计', self.show_fee_report),
            (' 生成月报', self.generate_monthly_report),
            (' 登出', self.logout)
        ]
        
        for text, slot in buttons_config:
            btn = QPushButton(text)
            btn.setStyleSheet(button_style)
            btn.clicked.connect(slot)
            left_layout.addWidget(btn)
        
        left_layout.addStretch()
        left_widget.setStyleSheet("background-color: #ecf0f1; border-radius: 10px;")
        return left_widget

    def create_right_panel(self):
        """创建右侧的摄像头和结果显示区域"""
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        recognition_group = QGroupBox('实时车牌识别')
        recognition_layout = QVBoxLayout()
        
        self.camera_label = QLabel("摄像头已关闭")
        self.camera_label.setMinimumSize(640, 480)
        self.camera_label.setAlignment(Qt.AlignCenter)
        self.camera_label.setStyleSheet("border: 2px dashed #bdc3c7; background-color: #ffffff; font-size: 20px; color: #7f8c8d;")
        recognition_layout.addWidget(self.camera_label)
        
        self.result_label = QLabel('识别结果将显示在这里')
        self.result_label.setAlignment(Qt.AlignCenter)
        self.result_label.setStyleSheet("font-size: 18px; color: #2c3e50; background-color: #ffffff; border-radius: 5px; padding: 10px; border: 1px solid #bdc3c7;")
        recognition_layout.addWidget(self.result_label)
        
        self.camera_btn = QPushButton('启动摄像头识别')
        self.camera_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 12px; font-size: 16px; font-weight: bold;")
        self.camera_btn.clicked.connect(self.toggle_camera)
        recognition_layout.addWidget(self.camera_btn)
        
        recognition_group.setLayout(recognition_layout)
        right_layout.addWidget(recognition_group)
        return right_widget

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        
    def show_add_user_dialog(self):
        dialog = AddUserDialog(self)
        dialog.exec_()
        
    def show_delete_user_dialog(self):
        dialog = DeleteUserDialog(self)
        dialog.exec_()
        
    def show_bind_vehicle_dialog(self):
        dialog = BindVehicleDialog(self)
        dialog.exec_()

    def show_search_dialog(self):
        dialog = SearchDialog(self)
        dialog.exec_()
        
    def show_fee_report(self):
        conn = sqlite3.connect('parking.db')
        cursor = conn.cursor()
        cursor.execute("SELECT DATE(entry_time), COUNT(*), SUM(fee) FROM parking_records WHERE fee IS NOT NULL GROUP BY DATE(entry_time) ORDER BY DATE(entry_time) DESC")
        results = cursor.fetchall()
        conn.close()
        
        report_text = "收费统计日报表\n" + "="*40 + "\n"
        report_text += "日期\t\t车辆数\t总收入(元)\n" + "-"*40 + "\n"
        
        total_income = 0
        for date, count, fee_sum in results:
            report_text += f"{date}\t{count}\t¥{fee_sum:.2f}\n"
            total_income += fee_sum
        
        report_text += "="*40 + f"\n总计收入: ¥{total_income:.2f}"
        QMessageBox.information(self, '收费统计报表', report_text)
        
    def toggle_camera(self):
        if not self.is_recognizing:
            try:
                self.recognizer.start_camera()
                self.timer.start(30)
                self.camera_btn.setText('停止识别')
                self.camera_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 12px; font-size: 16px; font-weight: bold;")
                self.is_recognizing = True
                self.result_label.setText('正在启动摄像头...')
            except Exception as e:
                QMessageBox.critical(self, "摄像头错误", f"无法启动摄像头: {e}")
        else:
            self.timer.stop()
            self.recognizer.stop_camera()
            self.camera_btn.setText('启动摄像头识别')
            self.camera_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 12px; font-size: 16px; font-weight: bold;")
            self.is_recognizing = False
            self.camera_label.setText("摄像头已关闭")
            self.camera_label.setStyleSheet("border: 2px dashed #bdc3c7; background-color: #ffffff; font-size: 20px; color: #7f8c8d;")
            self.result_label.setText('识别结果将显示在这里')
            
    def update_frame(self):
        frame, plate_number = self.recognizer.process_frame()
        if frame is not None:
            h, w, ch = frame.shape
            bytes_per_line = ch * w
            q_image = QImage(frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(q_image)
            self.camera_label.setPixmap(pixmap.scaled(self.camera_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
            
            if plate_number:
                self.result_label.setText(f'稳定识别结果: {plate_number}')
                self.handle_plate_recognition(plate_number)
            else:
                 self.result_label.setText('正在识别中...')
    
    def handle_plate_recognition(self, plate_number):
        self.toggle_camera()
        if self.parking.is_vehicle_inside(plate_number):
            exit_info = self.parking.vehicle_exit(plate_number)
            if exit_info:
                fee = exit_info["fee"]
                QMessageBox.information(self, '出场成功', f'车辆 {plate_number} 成功离场。\n应缴费用: ¥{fee:.2f}')
                self.show_exit_image()
        else:
            spot = self.parking.vehicle_entry(plate_number)
            if spot:
                QMessageBox.information(self, '入场成功', f'欢迎车辆 {plate_number}！\n已为您分配车位: {spot}号')
            else:
                QMessageBox.warning(self, '车位已满', '抱歉，当前停车场已无可用车位。')
    
    def show_exit_image(self):
        self.exit_window = QDialog(self)
        self.exit_window.setWindowTitle("感谢光临，请扫码支付")
        self.exit_window.setFixedSize(400, 300)
        label = QLabel(self.exit_window)
        pixmap = QPixmap("assets/zhifu.jpg")
        label.setPixmap(pixmap.scaled(400, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        label.setAlignment(Qt.AlignCenter)
        layout = QVBoxLayout()
        layout.addWidget(label)
        self.exit_window.setLayout(layout)
        self.exit_window.setWindowModality(Qt.ApplicationModal)
        self.exit_window.show()
        QTimer.singleShot(5000, self.exit_window.close)

    def logout(self):
        # <--- 修改在这里
        # 在需要时才导入 LoginWindow，打破循环
        from .login_window import LoginWindow
        
        self.close()
        self.login_window_instance = LoginWindow()
        self.login_window_instance.show()

    def generate_monthly_report(self):
        dialog = MonthSelectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            selected_date = dialog.get_selected_date()
            year, month = selected_date.year(), selected_date.month()
            self.create_excel_report(year, month)

    def create_excel_report(self, year, month):
        conn = sqlite3.connect('parking.db')
        cursor = conn.cursor()
        _, last_day = calendar.monthrange(year, month)
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{year}-{month:02d}-{last_day}"
        cursor.execute("""
            SELECT plate_number, COUNT(*), SUM(CASE WHEN exit_time IS NOT NULL THEN ROUND((julianday(exit_time) - julianday(entry_time)) * 24, 2) ELSE 0 END), SUM(COALESCE(fee, 0))
            FROM parking_records
            WHERE DATE(entry_time) BETWEEN ? AND ?
            GROUP BY plate_number
        """, (start_date, end_date))
        results = cursor.fetchall()
        conn.close()
        if not results:
            QMessageBox.information(self, "提示", f"{year}年{month}月无停车记录。")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月停车场月报"
        headers = ['序号', '车牌号', '停车次数', '总停车时长(小时)', '总费用(元)']
        ws.append(headers)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_align = Alignment(horizontal='center')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        total_fee_sum = 0
        for i, (plate, count, hours, fee) in enumerate(results, 1):
            ws.append([i, plate, count, f"{hours:.2f}", f"{fee:.2f}"])
            total_fee_sum += fee
        summary_row = len(results) + 3
        ws.cell(row=summary_row, column=1, value="月度总收入").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=f"¥{total_fee_sum:.2f}").font = Font(bold=True)
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col_letter].width = 20
        filename = f"停车场月报_{year}年{month:02d}月.xlsx"
        try:
            wb.save(filename)
            QMessageBox.information(self, '成功', f'月报已生成：{os.path.abspath(filename)}')
            os.startfile(os.path.abspath(filename))
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"无法保存Excel文件：{e}")
            
    def closeEvent(self, event):
        self.timer.stop()
        self.recognizer.stop_camera()
        event.accept()